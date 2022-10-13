import os
from os import listdir
from os.path import isfile, join
from openpyxl import load_workbook

# read all excel files in the current folder
files = [f for f in listdir('./') if isfile(join('./', f)) and f.split('.')[-1] == 'xlsx'] # TODO: more generalized path
os.mkdir('modified')
for file in files:
    workbook = load_workbook(filename=file)
    sheet = workbook['需求清單']

    # modify the R1 cell with 'Interdependency'
    if sheet['R1'].value == 'Independency':
        sheet['R1'] = 'Interdependency'
    
    # delete column N ~ Q
    if sheet['N3'].value == '最終須\n投入資源' and sheet['Q3'].value == '建議不納入範圍說明':
        # unmerge cells L1~Q1, L2~Q2, R1~U1, R2~U2
        sheet.unmerge_cells('L1:Q1')
        sheet.unmerge_cells('L2:Q2')
        sheet.unmerge_cells('R1:U1')
        sheet.unmerge_cells('R2:U2')

        sheet.delete_cols(14, 4)

        # unmerge cells L1~M1, L2~M2, N1~Q1, N2~Q2
        sheet.merge_cells('L1:M1')
        sheet.merge_cells('L2:M2')
        sheet.merge_cells('N1:Q1')
        sheet.merge_cells('N2:Q2')
        
    workbook.save(filename='modified/'+file)

